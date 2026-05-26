import streamlit as st
import sqlite3
import pandas as pd
from datetime import date
from io import BytesIO
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
# =====================================================
# STYLED EXCEL EXPORT FUNCTION
# =====================================================

def create_styled_excel(df, sheet_name="Report"):

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # COLORS
    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
        size=12
    )

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ADD DATAFRAME
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # STYLE HEADER
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # STYLE BODY
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # AUTO WIDTH
    for column_cells in ws.columns:

        length = 0
        column = column_cells[0].column_letter

        for cell in column_cells:
            try:
                if len(str(cell.value)) > length:
                    length = len(str(cell.value))
            except:
                pass

        ws.column_dimensions[column].width = length + 5

    # SAVE TO MEMORY
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
# ==========================================
# FLEXWOLF PREMIUM LIGHT THEME
# ==========================================

st.set_page_config(
    page_title="FLEXWOLF ERP",
    page_icon="🐺",
    layout="wide"
)

st.markdown(
    """
    <style>

    /* MAIN APP */
    .stApp {
        background-color: #F3F4F6;
        color: #111827;
    }

    /* SIDEBAR */
    section[data-testid="stSidebar"] {
        background-color: #1F4E78;
    }

    /* SIDEBAR TEXT */
    section[data-testid="stSidebar"] * {
        color: white !important;
    }

    /* HEADINGS */
    h1, h2, h3 {
        color: #1F4E78 !important;
        font-weight: bold;
    }

    /* BUTTONS */
    .stButton>button {

        background-color: #2563EB;

        color: white;

        border-radius: 10px;

        border: none;

        padding: 10px 20px;

        font-weight: bold;
    }

    .stButton>button:hover {

        background-color: #1D4ED8;

        color: white;
    }

    /* INPUTS */
    .stTextInput input,
    .stNumberInput input,
    .stDateInput input,
    .stSelectbox div[data-baseweb="select"] {

        background-color: white !important;

        color: black !important;

        border-radius: 10px !important;

        border: 1px solid #D1D5DB !important;
    }

    /* METRIC CARDS */
    div[data-testid="metric-container"] {

        background-color: white;

        border-radius: 15px;

        padding: 15px;

        border: 1px solid #E5E7EB;

        box-shadow: 0px 2px 8px rgba(0,0,0,0.08);
    }

    /* DATAFRAME */
    .stDataFrame {

        background-color: white;

        border-radius: 10px;

        padding: 10px;

        border: 1px solid #E5E7EB;
    }

    /* SUCCESS */
    .stSuccess {

        border-radius: 10px;
    }

    /* TABLE */
    table {

        border-radius: 10px;

        overflow: hidden;
    }

    </style>
    """,
    unsafe_allow_html=True
)
# LOGIN SESSION
if "logged_in" not in st.session_state:

    st.session_state.logged_in = False


# PAGE CONFIG
st.set_page_config(
    page_title="Garment ERP",
    page_icon="🏭",
    layout="wide"
)

# DATABASE
conn = sqlite3.connect(
    "garment_factory.db",
    check_same_thread=False
)

cursor = conn.cursor()

# TITLE

# ==========================================
# LOGIN SYSTEM
# ==========================================

if not st.session_state.logged_in:

    st.title("🔐 FLEXWOLF ERP LOGIN")

    username = st.text_input(
        "Username"
    )

    password = st.text_input(
        "Password",
        type="password"
    )

    if st.button("Login"):

        user = pd.read_sql(
            f"""
            SELECT *
            FROM users
            WHERE username='{username}'
            AND password='{password}'
            """,
            conn
        )

        if len(user) > 0:

            st.session_state.logged_in = True
            st.session_state.role = (
    user.iloc[0]["role"]
    )

            st.success(
                "Login Successful"
            )

            st.rerun()

        else:

            st.error(
                "Invalid Username or Password"
            )

    st.stop()
# ==========================================
# ROLE BASED MENU
# ==========================================

role = st.session_state.role

# =====================================================
# SIDEBAR
# =====================================================

st.sidebar.title("Select Module")

main_menu = st.sidebar.radio(
    "Main Menu",
    [
        "Dashboard",
        "Karigar Management",
        "Staff Management",
        "Vendor Management",
        "Inventory",
        "Settings"
    ]
)

# =====================================================
# KARIGAR MANAGEMENT
# =====================================================

if main_menu == "Karigar Management":

    menu = st.sidebar.radio(
        "Karigar Menu",
        [
            "Karigar Master",
            "Machine Master",
            "Article Master",
            "Rate Master",
            "Production",
            "Advance Entry",
            "Karigar Ledger"
        ]
    )

# =====================================================
# STAFF MANAGEMENT
# =====================================================

elif main_menu == "Staff Management":

    menu = st.sidebar.radio(
        "Staff Menu",
        [
            "Staff Master",
            "Staff Advance",
            "Staff Salary"
        ]
    )

# =====================================================
# VENDOR MANAGEMENT
# =====================================================

elif main_menu == "Vendor Management":

    menu = st.sidebar.radio(
        "Vendor Menu",
        [
            "Vendors"
        ]
    )

# =====================================================
# INVENTORY
# =====================================================

elif main_menu == "Inventory":

    menu = st.sidebar.radio(
        "Inventory Menu",
        [
            "Fabric Stock",
            "Cutting",
            "Box Packing"
        ]
    )

# =====================================================
# SETTINGS
# =====================================================

elif main_menu == "Settings":

    menu = st.sidebar.radio(
        "Settings Menu",
        [
            "Manage Data",
            "Export Reports",
            "User Management"
        ]
    )

# =====================================================
# DASHBOARD
# =====================================================

else:

    menu = "Dashboard"

# ==========================================
# DASHBOARD
# ==========================================

if menu == "Dashboard":
    st.image(
    "flexwolf_logo.png",
    width=180
)

    st.title("FLEXWOLF")

    # ======================================
    # TOTAL PRODUCTION
    # ======================================

    production_total = pd.read_sql(
        """
        SELECT SUM(amount) as total
        FROM production
        """,
        conn
    )

    total_production = (
        production_total.iloc[0]["total"]
    )

    if total_production is None:
        total_production = 0

    # ======================================
    # TOTAL ADVANCE
    # ======================================

    advance_total = pd.read_sql(
        """
        SELECT SUM(amount) as total
        FROM advances
        """,
        conn
    )

    total_advance = (
        advance_total.iloc[0]["total"]
    )

    if total_advance is None:
        total_advance = 0

    # ======================================
    # TOTAL CUTTING
    # ======================================

    cutting_total = pd.read_sql(
        """
        SELECT SUM(total_pieces) as total
        FROM cutting
        """,
        conn
    )

    total_cutting = (
        cutting_total.iloc[0]["total"]
    )

    if total_cutting is None:
        total_cutting = 0

    # ======================================
    # TOTAL PACKING
    # ======================================

    packing_total = pd.read_sql(
        """
        SELECT SUM(total_pieces) as total
        FROM box_packing
        """,
        conn
    )

    total_packing = (
        packing_total.iloc[0]["total"]
    )

    if total_packing is None:
        total_packing = 0

    # ======================================
    # BALANCE PAYABLE
    # ======================================

    balance_payable = (
        total_production -
        total_advance
    )

    # ======================================
    # TOP SUMMARY
    # ======================================

    col1, col2, col3, col4 = st.columns(4)

    col1.metric(
        "Production Value",
        f"₹ {total_production}"
    )

    col2.metric(
        "Advance Paid",
        f"₹ {total_advance}"
    )

    col3.metric(
        "Cutting PCS",
        total_cutting
    )

    col4.metric(
        "Packing PCS",
        total_packing
    )

    st.divider()

    # ======================================
    # BALANCE BOX
    # ======================================

    st.success(
        f"💰 Total Balance Payable: ₹ {balance_payable}"
    )

    st.divider()

    # ======================================
    # TOP KARIGARS
    # ======================================

    st.subheader("🏆 Top Karigars")

    top_karigar = pd.read_sql(
        """
        SELECT

            karigar_name,
            SUM(amount) as total_amount

        FROM production

        GROUP BY karigar_name

        ORDER BY total_amount DESC
        """,
        conn
    )

    st.dataframe(
        top_karigar,
        use_container_width=True
    )

    st.divider()

    # ======================================
    # LIVE FABRIC STOCK
    # ======================================

    st.subheader("🧵 Live Fabric Stock")

    fabric_stock = pd.read_sql(
        """
        SELECT

            fabric_name,
            color,

            SUM(total_rolls) as rolls,

            SUM(total_weight) as weight

        FROM fabric_stock

        GROUP BY fabric_name, color
        """,
        conn
    )

    st.dataframe(
        fabric_stock,
        use_container_width=True
    )

    st.divider()

    # ======================================
    # RECENT PRODUCTION
    # ======================================

    st.subheader("📦 Recent Production")

    recent_production = pd.read_sql(
        """
        SELECT *

        FROM production

        ORDER BY id DESC

        LIMIT 10
        """,
        conn
    )

    st.dataframe(
        recent_production,
        use_container_width=True
    )

# MACHINE MASTER
elif menu == "Machine Master":

    st.header("Machine Master")

    machine = st.text_input(
        "Machine Name"
    )

    if st.button("Save Machine"):

        cursor.execute(
            """
            INSERT INTO machines(machine_name)
            VALUES(?)
            """,
            (machine,)
        )

        conn.commit()

        st.success("Machine Saved")

    data = pd.read_sql(
        "SELECT * FROM machines",
        conn
    )

    st.dataframe(data)
    # KARIGAR MASTER
elif menu == "Karigar Master":

    st.header("Karigar Master")

    machine_data = pd.read_sql(
        "SELECT * FROM machines",
        conn
    )

    if len(machine_data) > 0:

        karigar = st.text_input(
            "Karigar Name"
        )

        machine = st.selectbox(
            "Select Machine",
            machine_data["machine_name"]
        )

        if st.button("Save Karigar"):

            cursor.execute(
                """
                INSERT INTO karigars
                (
                    karigar_name,
                    machine_name
                )
                VALUES(?,?)
                """,
                (
                    karigar,
                    machine
                )
            )

            conn.commit()

            st.success("Karigar Saved")

        data = pd.read_sql(
            "SELECT * FROM karigars",
            conn
        )

        st.dataframe(data)

    else:

        st.warning(
            "Please add machine first"
        )
        # =====================================
# KARIGAR LIST
# =====================================

st.subheader("Karigar List")

karigar_df = pd.read_sql(
    """
    SELECT *
    FROM karigars
    ORDER BY karigar_name
    """,
    conn
)

st.dataframe(
    karigar_df,
    use_container_width=True
)

# =====================================
# DELETE KARIGAR
# =====================================

st.subheader("🗑 Delete Karigar")

if not karigar_df.empty:

    delete_karigar = st.selectbox(
        "Select Karigar",
        karigar_df["karigar_name"]
    )

    if st.button("Delete Karigar"):

        # DELETE KARIGAR
        cursor.execute(
            """
            DELETE FROM karigars
            WHERE karigar_name=?
            """,
            (delete_karigar,)
        )

        # DELETE ADVANCES
        cursor.execute(
            """
            DELETE FROM advances
            WHERE karigar_name=?
            """,
            (delete_karigar,)
        )

        # DELETE PRODUCTION
        cursor.execute(
            """
            DELETE FROM production
            WHERE karigar_name=?
            """,
            (delete_karigar,)
        )

        conn.commit()

        st.success("Karigar Deleted Successfully")

        st.rerun()
        # ARTICLE MASTER
elif menu == "Article Master":

    st.header("Article Master")

    article = st.text_input(
        "Article Name"
    )

    if st.button("Save Article"):

        cursor.execute(
            """
            INSERT INTO articles(article_name)
            VALUES(?)
            """,
            (article,)
        )

        conn.commit()

        st.success("Article Saved")

    data = pd.read_sql(
        "SELECT * FROM articles",
        conn
    )

    st.dataframe(data)
    # RATE MASTER
elif menu == "Rate Master":

    st.header("Rate Master")

    article_data = pd.read_sql(
        "SELECT * FROM articles",
        conn
    )

    machine_data = pd.read_sql(
        "SELECT * FROM machines",
        conn
    )

    if len(article_data) > 0 and len(machine_data) > 0:

        article = st.selectbox(
            "Select Article",
            article_data["article_name"]
        )

        machine = st.selectbox(
            "Select Machine",
            machine_data["machine_name"]
        )

        rate = st.number_input(
            "Piece Rate",
            min_value=0.0
        )

        if st.button("Save Rate"):

            cursor.execute(
                """
                INSERT INTO article_rates
                (
                    article_name,
                    machine_name,
                    rate
                )
                VALUES(?,?,?)
                """,
                (
                    article,
                    machine,
                    rate
                )
            )

            conn.commit()

            st.success("Rate Saved")

        data = pd.read_sql(
            "SELECT * FROM article_rates",
            conn
        )

        st.dataframe(data)

    else:

        st.warning(
            "Please add articles and machines first"
        )
            # ==========================================
    # EDIT RATE
    # ==========================================

    st.divider()

    st.subheader("✏️ Edit Rate")

    rate_data = pd.read_sql(
        """
        SELECT *
        FROM article_rates
        """,
        conn
    )

    st.dataframe(
        rate_data,
        use_container_width=True
    )

    edit_id = st.number_input(
        "Enter Rate ID To Edit",
        min_value=1,
        key="edit_rate_id"
    )

    new_rate = st.number_input(
        "Enter New Rate",
        min_value=0.0,
        key="new_rate_value"
    )

    if st.button(
        "Update Rate",
        key="update_rate_btn"
    ):

        cursor.execute(
            """
            UPDATE article_rates
            SET rate=?
            WHERE id=?
            """,
            (
                new_rate,
                edit_id
            )
        )

        conn.commit()

        st.success(
            "Rate Updated Successfully"
        )

# =========================================
# PRODUCTION MODULE
# =========================================

elif menu == "Production":

    st.title("🏭 Production Entry")

    # =========================
    # LOAD KARIGARS
    # =========================

    karigar_df = pd.read_sql(
        "SELECT karigar_name FROM karigars ORDER BY karigar_name",
        conn
    )

    karigar_list = karigar_df["karigar_name"].tolist()

    # =========================
    # LOAD MACHINES
    # =========================

    machine_df = pd.read_sql(
        "SELECT machine_name FROM machines ORDER BY machine_name",
        conn
    )

    machine_list = machine_df["machine_name"].tolist()

    # =========================
    # LOAD ARTICLES
    # =========================

    article_df = pd.read_sql(
        "SELECT article_name FROM articles ORDER BY article_name",
        conn
    )

    article_list = article_df["article_name"].tolist()

    # =========================
    # ENTRY FORM
    # =========================

    entry_date = st.date_input(
        "Entry Date",
        date.today()
    )

    karigar_name = st.selectbox(
        "Select Karigar",
        karigar_list
    )

    machine_name = st.selectbox(
        "Select Machine",
        machine_list
    )

    article_name = st.selectbox(
        "Select Article",
        article_list
    )

    qty = st.number_input(
        "Pieces Qty",
        min_value=0,
        step=1
    )

    # =========================
    # FETCH RATE
    # =========================

    rate_df = pd.read_sql(
        """
        SELECT rate
        FROM article_rates
        WHERE article_name=?
        AND machine_name=?
        """,
        conn,
        params=(
            article_name,
            machine_name
        )
    )

    if not rate_df.empty:

        rate = rate_df.iloc[0]["rate"]

    else:

        rate = 0

        st.warning("⚠ Rate Not Found For Selected Machine")

    # =========================
    # CALCULATE AMOUNT
    # =========================

    amount = qty * rate

    col1, col2 = st.columns(2)

    with col1:

        st.info(f"Rate : ₹ {rate}")

    with col2:

        st.success(f"Amount : ₹ {amount}")

    # =========================
    # SAVE BUTTON
    # =========================

    if st.button("Save Production"):

        cursor.execute(
            """
            INSERT INTO production(
                entry_date,
                karigar_name,
                machine_name,
                article_name,
                qty,
                rate,
                amount,
                payment_status
            )
            VALUES(?,?,?,?,?,?,?,?)
            """,
            (
                str(entry_date),
                karigar_name,
                machine_name,
                article_name,
                qty,
                rate,
                amount,
                "Pending"
            )
        )

        conn.commit()

        st.success("✅ Production Saved")

    st.divider()

    # =====================================
    # KARIGAR PAYMENT SUMMARY
    # =====================================

    st.subheader("💰 Karigar Payment Summary")

    summary_df = pd.read_sql(
        """
        SELECT
            karigar_name,
            SUM(amount) as total_production
        FROM production
        GROUP BY karigar_name
        """,
        conn
    )

    for index, row in summary_df.iterrows():

        karigar = row["karigar_name"]

        total_prod = row["total_production"]

        # =========================
        # TOTAL ADVANCE
        # =========================

        adv_df = pd.read_sql(
            """
            SELECT SUM(amount) as total_advance
            FROM advances
            WHERE karigar_name=?
            """,
            conn,
            params=(karigar,)
        )

        total_adv = adv_df.iloc[0]["total_advance"]

        if total_adv is None:

            total_adv = 0

        # =========================
        # FINAL PAYMENT
        # =========================

        final_payment = total_prod - total_adv

        # =========================
        # PAYMENT STATUS
        # =========================

        status_df = pd.read_sql(
            """
            SELECT payment_status
            FROM production
            WHERE karigar_name=?
            LIMIT 1
            """,
            conn,
            params=(karigar,)
        )

        if not status_df.empty:

            payment_status = status_df.iloc[0]["payment_status"]

        else:

            payment_status = "Pending"

        # =========================
        # UI CARD
        # =========================

        st.markdown("---")

        col1, col2, col3, col4, col5 = st.columns(5)

        with col1:

            st.write(f"👨‍🔧 {karigar}")

        with col2:

            st.write(f"Production : ₹ {total_prod}")

        with col3:

            st.write(f"Advance : ₹ {total_adv}")

        with col4:

            st.write(f"Final : ₹ {final_payment}")

        with col5:

            if payment_status == "Pending":

                if st.button(
                    f"✅ Mark Paid - {karigar}",
                    key=f"paid_{karigar}"
                ):

                    cursor.execute(
                        """
                        UPDATE production
                        SET payment_status='Paid'
                        WHERE karigar_name=?
                        """,
                        (karigar,)
                    )

                    conn.commit()

                    st.rerun()

            else:

                st.success("Paid")

    # =====================================
    # PRODUCTION HISTORY
    # =====================================

    st.divider()

    st.subheader("📋 Production History")

    history_df = pd.read_sql(
        """
        SELECT *
        FROM production
        ORDER BY id DESC
        """,
        conn
    )

    st.dataframe(
        history_df,
        use_container_width=True
    )

    # =====================================
    # DELETE ENTRY
    # =====================================

    if not history_df.empty:

        selected_id = st.selectbox(
            "Select Production ID To Delete",
            history_df["id"]
        )

        if st.button("🗑 Delete Production Entry"):

            cursor.execute(
                """
                DELETE FROM production
                WHERE id=?
                """,
                (selected_id,)
            )

            conn.commit()

            st.success("Production Entry Deleted")

            st.rerun()

    # FABRIC STOCK
elif menu == "Fabric Stock":

    st.header("Fabric Inward")

    entry_date = st.date_input(
        "Entry Date",
        date.today()
    )

    fabric_name = st.text_input(
        "Fabric Name"
    )

    color = st.text_input(
        "Color"
    )

    total_rolls = st.number_input(
        "Number Of Rolls",
        min_value=1
    )

    total_weight = st.number_input(
        "Total Weight (KG)",
        min_value=0.0
    )

    if st.button("Save Fabric"):

        cursor.execute(
            """
            INSERT INTO fabric_stock
            (
                entry_date,
                fabric_name,
                color,
                total_rolls,
                total_weight
            )
            VALUES(?,?,?,?,?)
            """,
            (
                str(entry_date),
                fabric_name,
                color,
                total_rolls,
                total_weight
            )
        )

        conn.commit()

        st.success("Fabric Saved")

    st.subheader("Fabric History")

    data = pd.read_sql(
        """
        SELECT *
        FROM fabric_stock
        ORDER BY id DESC
        """,
        conn
    )

    st.dataframe(data)

    st.subheader("Live Fabric Stock")

    stock = pd.read_sql(
        """
        SELECT
            fabric_name,
            color,
            SUM(total_rolls) as stock_rolls,
            SUM(total_weight) as stock_weight
        FROM fabric_stock
        GROUP BY fabric_name, color
        """,
        conn
    )

    st.dataframe(stock)
   # CUTTING
elif menu == "Cutting":

    st.header("✂️ Cutting Entry")

    # FABRIC STOCK
    stock_data = pd.read_sql(
        """
        SELECT
            fabric_name,
            color,
            SUM(total_rolls) as stock_rolls
        FROM fabric_stock
        GROUP BY fabric_name, color
        """,
        conn
    )

    if len(stock_data) > 0:

        # DISPLAY NAME
        stock_data["display"] = (
            stock_data["fabric_name"]
            + " - "
            + stock_data["color"]
        )

        # FABRIC SELECT
        selected = st.selectbox(
            "Select Fabric Color",
            stock_data["display"]
        )

        # ROW
        row = stock_data[
            stock_data["display"] == selected
        ]

        fabric_name = row.iloc[0]["fabric_name"]
        color = row.iloc[0]["color"]

        current_stock = int(
            row.iloc[0]["stock_rolls"]
        )

        st.success(
            f"Current Stock: {current_stock} Rolls"
        )

        # LOT NUMBER
        lot_no = st.text_input(
            "Lot Number"
        )

        # USED ROLLS
        used_rolls = st.number_input(
            "Used Rolls",
            min_value=1
        )

        st.subheader("Size Wise Cutting")

        # SIZE INPUTS
        col1, col2, col3, col4 = st.columns(4)

        s_qty = col1.number_input(
            "S",
            min_value=0
        )

        m_qty = col2.number_input(
            "M",
            min_value=0
        )

        l_qty = col3.number_input(
            "L",
            min_value=0
        )

        xl_qty = col4.number_input(
            "XL",
            min_value=0
        )

        col5, col6, col7 = st.columns(3)

        xxl_qty = col5.number_input(
            "XXL",
            min_value=0
        )

        xxxl_qty = col6.number_input(
            "3XL",
            min_value=0
        )

        xxxxl_qty = col7.number_input(
            "4XL",
            min_value=0
        )

        # TOTAL PCS
        total_pieces = (
            s_qty +
            m_qty +
            l_qty +
            xl_qty +
            xxl_qty +
            xxxl_qty +
            xxxxl_qty
        )

        st.info(
            f"Total Pieces: {total_pieces}"
        )

        # BALANCE
        balance = current_stock - used_rolls

        st.warning(
            f"Balance Stock: {balance} Rolls"
        )

        # SAVE
        if st.button("Save Cutting"):

            # STOCK DEDUCTION
            cursor.execute(
                """
                INSERT INTO fabric_stock
                (
                    entry_date,
                    fabric_name,
                    color,
                    total_rolls,
                    total_weight
                )
                VALUES(?,?,?,?,?)
                """,
                (
                    str(date.today()),
                    fabric_name,
                    color,
                    -used_rolls,
                    0
                )
            )

            # CUTTING SAVE
            cursor.execute(
                """
                INSERT INTO cutting
                (
                    entry_date,
                    fabric_name,
                    color,
                    lot_no,
                    used_rolls,

                    s_qty,
                    m_qty,
                    l_qty,
                    xl_qty,
                    xxl_qty,
                    xxxl_qty,
                    xxxxl_qty,

                    total_pieces
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    str(date.today()),
                    fabric_name,
                    color,
                    lot_no,
                    used_rolls,

                    s_qty,
                    m_qty,
                    l_qty,
                    xl_qty,
                    xxl_qty,
                    xxxl_qty,
                    xxxxl_qty,

                    total_pieces
                )
            )

            conn.commit()

            st.success(
                "Cutting Saved Successfully"
            )

    else:

        st.warning(
            "No Fabric Stock Available"
        )

    # HISTORY
    st.subheader("Cutting History")

    cutting_data = pd.read_sql(
        """
        SELECT *
        FROM cutting
        ORDER BY id DESC
        """,
        conn
    )

    st.dataframe(
        cutting_data,
        use_container_width=True
    )
   # ==========================================
# BOX PACKING
# ==========================================

if menu == "Box Packing":

    st.header("📦 Box Packing")

    article_data = pd.read_sql(
        "SELECT * FROM articles",
        conn
    )

    if len(article_data) > 0:

        packing_date = st.date_input(
            "Packing Date",
            date.today()
        )

        article = st.selectbox(
            "Select Article",
            article_data["article_name"]
        )

        color = st.text_input("Color")

        box_number = st.text_input(
            "Box Number"
        )

        st.subheader("Size Wise Packing")

        col1, col2, col3, col4 = st.columns(4)

        s_qty = col1.number_input("S", min_value=0)

        m_qty = col2.number_input("M", min_value=0)

        l_qty = col3.number_input("L", min_value=0)

        xl_qty = col4.number_input("XL", min_value=0)

        col5, col6, col7 = st.columns(3)

        xxl_qty = col5.number_input("XXL", min_value=0)

        xxxl_qty = col6.number_input("3XL", min_value=0)

        xxxxl_qty = col7.number_input("4XL", min_value=0)

        total_pieces = (
            s_qty +
            m_qty +
            l_qty +
            xl_qty +
            xxl_qty +
            xxxl_qty +
            xxxxl_qty
        )

        st.info(
            f"Total Pieces: {total_pieces}"
        )

        if st.button("Save Packing"):

            cursor.execute(
                """
                INSERT INTO box_packing
                (
                    packing_date,
                    article_name,
                    color,
                    box_number,

                    s_qty,
                    m_qty,
                    l_qty,
                    xl_qty,
                    xxl_qty,
                    xxxl_qty,
                    xxxxl_qty,

                    total_pieces
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    str(packing_date),
                    article,
                    color,
                    box_number,

                    s_qty,
                    m_qty,
                    l_qty,
                    xl_qty,
                    xxl_qty,
                    xxxl_qty,
                    xxxxl_qty,

                    total_pieces
                )
            )

            conn.commit()

            st.success(
                "Packing Saved Successfully"
            )

    # HISTORY
    st.subheader("Packing History")

    packing_data = pd.read_sql(
        """
        SELECT

            box_number as 'Box No',

            article_name as 'Article',

            color as 'Color',

            s_qty as 'S',

            m_qty as 'M',

            l_qty as 'L',

            xl_qty as 'XL',

            xxl_qty as 'XXL',

            xxxl_qty as '3XL',

            xxxxl_qty as '4XL',

            total_pieces as 'Total PCS',

            packing_date as 'Date'

        FROM box_packing

        ORDER BY id DESC
        """,
        conn
    )

    st.dataframe(
        packing_data,
        use_container_width=True
    )

# ==========================================
# ADVANCE ENTRY
# ==========================================

if menu == "Advance Entry":

    st.header("💰 Advance Entry")

    karigar_data = pd.read_sql(
        "SELECT * FROM karigars",
        conn
    )

    if len(karigar_data) > 0:

        karigar = st.selectbox(
            "Select Karigar",
            karigar_data["karigar_name"]
        )

        entry_date = st.date_input(
            "Date",
            date.today()
        )

        amount = st.number_input(
            "Advance Amount",
            min_value=0.0
        )

        remarks = st.text_input(
            "Remarks"
        )

        if st.button("Save Advance"):

            cursor.execute(
                """
                INSERT INTO advances
                (
                    entry_date,
                    karigar_name,
                    amount,
                    remarks
                )
                VALUES(?,?,?,?)
                """,
                (
                    str(entry_date),
                    karigar,
                    amount,
                    remarks
                )
            )

            conn.commit()

            st.success(
                "Advance Saved Successfully"
            )

    st.subheader("Advance History")

    advance_history = pd.read_sql(
        """
        SELECT

            entry_date as 'Date',

            karigar_name as 'Karigar',

            amount as 'Amount',

            remarks as 'Remarks'

        FROM advances

        ORDER BY id DESC
        """,
        conn
    )

    st.dataframe(
        advance_history,
        use_container_width=True
    )
# ==========================================
# ADVANCE ENTRY
# ==========================================

if menu == "Advance Entry":

    st.header("💰 Advance Entry")

    karigar_df = pd.read_sql(
        "SELECT * FROM karigars",
        conn
    )

    if len(karigar_df) > 0:

        karigar_name = st.selectbox(
            "Select Karigar",
            karigar_df["karigar_name"],
            key="advance_karigar"
        )

        advance_date = st.date_input(
            "Advance Date",
            date.today(),
            key="advance_date"
        )

        advance_amount = st.number_input(
            "Advance Amount",
            min_value=0.0,
            key="advance_amount"
        )

        advance_remarks = st.text_input(
            "Remarks",
            key="advance_remarks"
        )

        if st.button(
            "Save Advance",
            key="save_advance_btn"
        ):

            cursor.execute(
                """
                INSERT INTO advances
                (
                    entry_date,
                    karigar_name,
                    amount,
                    remarks
                )
                VALUES(?,?,?,?)
                """,
                (
                    str(advance_date),
                    karigar_name,
                    advance_amount,
                    advance_remarks
                )
            )

            conn.commit()

            st.success(
                "Advance Saved Successfully"
            )

    else:

        st.warning(
            "Please Add Karigar First"
        )

    # ======================================
    # ADVANCE HISTORY
    # ======================================

    st.subheader("Advance History")

    advance_history = pd.read_sql(
        """
        SELECT

            entry_date as 'Date',

            karigar_name as 'Karigar',

            amount as 'Amount',

            remarks as 'Remarks'

        FROM advances

        ORDER BY id DESC
        """,
        conn
    )

    st.dataframe(
        advance_history,
        use_container_width=True
    )

# =====================================
# KARIGAR LEDGER
# =====================================

if menu == "Karigar Ledger":

    st.header("📒 Karigar Ledger")

    karigar_data = pd.read_sql(
        "SELECT * FROM karigars",
        conn
    )

    if len(karigar_data) > 0:

        karigar = st.selectbox(
            "Select Karigar",
            karigar_data["karigar_name"]
        )

        # =====================================================
        # PRODUCTION HISTORY
        # =====================================================

    st.divider()

    st.subheader("📜 Production History")

    production_history = pd.read_sql(
    """
    SELECT *
    FROM production
    ORDER BY id DESC
    """,
    conn
    )

    if len(production_history) > 0:

      for index, row in production_history.iterrows():

        with st.expander(
           f"{row['karigar_name']} | {row['machine_name']} | ₹ {row['amount']} | {row['payment_status']}"
        ):

            st.write(f"Date: {row['entry_date']}")
            st.write(f"Article: {row['article_name']}")
            st.write(f"Qty: {row['qty']}")
            st.write(f"Rate: ₹ {row['rate']}")
            st.write(f"Amount: ₹ {row['amount']}")

            # EDIT SECTION
            new_qty = st.number_input(
                "Edit Qty",
                value=float(row["qty"]),
                key=f"qty_{row['id']}"
            )

            new_rate = st.number_input(
                "Edit Rate",
                value=float(row["rate"]),
                key=f"rate_{row['id']}"
            )

            new_amount = new_qty * new_rate

            st.info(f"Updated Amount = ₹ {new_amount}")

            col1, col2, col3 = st.columns(3)

            # UPDATE
            with col1:

                if st.button(
                    "Update",
                    key=f"update_prod_{row['id']}"
                ):

                    cursor.execute(
                        """
                        UPDATE production
                        SET qty=?,
                            rate=?,
                            amount=?
                        WHERE id=?
                        """,

                        (
                            new_qty,
                            new_rate,
                            new_amount,
                            row["id"]
                        )
                    )

                    conn.commit()

                    st.success("Updated Successfully")

                    st.rerun()

            # PAID BUTTON
            with col2:

                if row["payment_status"] == "Pending":

                    if st.button(
                        "Mark Paid",
                        key=f"paid_prod_{row['id']}"
                    ):

                        cursor.execute(
                            """
                            UPDATE production
                            SET payment_status='Paid'
                            WHERE id=?
                            """,
                            (row["id"],)
                        )

                        conn.commit()

                        st.success("Payment Marked Paid")

                        st.rerun()

            # DELETE
            with col3:

                if st.button(
                    "Delete",
                    key=f"delete_prod_{row['id']}"
                ):

                    cursor.execute(
                        "DELETE FROM production WHERE id=?",
                        (row["id"],)
                    )

                    conn.commit()

                    st.success("Deleted Successfully")

                    st.rerun()
                    
        # ==========================================
# MANAGE DATA
# ==========================================

if menu == "Manage Data":

    st.header("🗑️ Manage Data")

    table_option = st.selectbox(
        "Select Module",
        [
            "Production",
            "Fabric Stock",
            "Cutting",
            "Box Packing",
            "Advances"
        ]
    )

    # ======================================
    # PRODUCTION
    # ======================================

    if table_option == "Production":

        data = pd.read_sql(
            """
            SELECT *
            FROM production
            ORDER BY id DESC
            """,
            conn
        )

        st.dataframe(
            data,
            use_container_width=True
        )

        delete_id = st.number_input(
            "Enter Production ID To Delete",
            min_value=1
        )

        if st.button("Delete Production"):

            cursor.execute(
                """
                DELETE FROM production
                WHERE id=?
                """,
                (delete_id,)
            )

            conn.commit()

            st.success(
                "Production Deleted"
            )

    # ======================================
    # FABRIC STOCK
    # ======================================

    elif table_option == "Fabric Stock":

        data = pd.read_sql(
            """
            SELECT *
            FROM fabric_stock
            ORDER BY id DESC
            """,
            conn
        )

        st.dataframe(
            data,
            use_container_width=True
        )

        delete_id = st.number_input(
            "Enter Fabric Stock ID To Delete",
            min_value=1
        )

        if st.button("Delete Fabric Stock"):

            cursor.execute(
                """
                DELETE FROM fabric_stock
                WHERE id=?
                """,
                (delete_id,)
            )

            conn.commit()

            st.success(
                "Fabric Stock Deleted"
            )

    # ======================================
    # CUTTING
    # ======================================

    elif table_option == "Cutting":

        data = pd.read_sql(
            """
            SELECT *
            FROM cutting
            ORDER BY id DESC
            """,
            conn
        )

        st.dataframe(
            data,
            use_container_width=True
        )

        delete_id = st.number_input(
            "Enter Cutting ID To Delete",
            min_value=1
        )

        if st.button("Delete Cutting"):

            cursor.execute(
                """
                DELETE FROM cutting
                WHERE id=?
                """,
                (delete_id,)
            )

            conn.commit()

            st.success(
                "Cutting Deleted"
            )

    # ======================================
    # BOX PACKING
    # ======================================

    elif table_option == "Box Packing":

        data = pd.read_sql(
            """
            SELECT *
            FROM box_packing
            ORDER BY id DESC
            """,
            conn
        )

        st.dataframe(
            data,
            use_container_width=True
        )

        delete_id = st.number_input(
            "Enter Packing ID To Delete",
            min_value=1
        )

        if st.button("Delete Packing"):

            cursor.execute(
                """
                DELETE FROM box_packing
                WHERE id=?
                """,
                (delete_id,)
            )

            conn.commit()

            st.success(
                "Packing Deleted"
            )

    # ======================================
    # ADVANCES
    # ======================================

    elif table_option == "Advances":

        data = pd.read_sql(
            """
            SELECT *
            FROM advances
            ORDER BY id DESC
            """,
            conn
        )

        st.dataframe(
            data,
            use_container_width=True
        )

        delete_id = st.number_input(
            "Enter Advance ID To Delete",
            min_value=1
        )

        if st.button("Delete Advance"):

            cursor.execute(
                """
                DELETE FROM advances
                WHERE id=?
                """,
                (delete_id,)
            )

            conn.commit()

            st.success(
                "Advance Deleted"
            )

# =====================================================
# EXPORT REPORTS MODULE
# =====================================================

elif menu == "Export Reports":

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Alignment
    from openpyxl.styles import Border
    from openpyxl.styles import Side

    st.header("📤 Export Reports")

    # =================================================
    # EXCEL FUNCTION
    # =================================================

    def create_excel(df, sheet_name):

        wb = Workbook()

        ws = wb.active

        ws.title = sheet_name

        # =============================================
        # HEADER STYLE
        # =============================================

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        header_font = Font(
            color="FFFFFF",
            bold=True
        )

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # =============================================
        # HEADERS
        # =============================================

        for col_num, column_name in enumerate(df.columns, 1):

            cell = ws.cell(
                row=1,
                column=col_num
            )

            cell.value = column_name

            cell.fill = header_fill

            cell.font = header_font

            cell.alignment = Alignment(
                horizontal="center"
            )

            cell.border = border

        # =============================================
        # DATA
        # =============================================

        for row_num, row_data in enumerate(df.values, 2):

            for col_num, value in enumerate(row_data, 1):

                cell = ws.cell(
                    row=row_num,
                    column=col_num
                )

                cell.value = value

                cell.border = border

                cell.alignment = Alignment(
                    horizontal="center"
                )

        # =============================================
        # AUTO WIDTH
        # =============================================

        for column_cells in ws.columns:

            max_length = 0

            column = column_cells[0].column_letter

            for cell in column_cells:

                try:

                    if len(str(cell.value)) > max_length:

                        max_length = len(str(cell.value))

                except:
                    pass

            adjusted_width = max_length + 5

            ws.column_dimensions[column].width = adjusted_width

        # =============================================
        # SAVE
        # =============================================

        output = BytesIO()

        wb.save(output)

        output.seek(0)

        return output

    # =================================================
    # REPORT OPTIONS
    # =================================================

    report_type = st.selectbox(
        "Select Report",
        [
            "Production Report",
            "Karigar Ledger",
            "Vendor Bills",
            "Vendor Payments",
            "Staff Salary",
            "Staff Advance"
        ]
    )

    # =================================================
    # PRODUCTION REPORT
    # =================================================

    if report_type == "Production Report":

        df = pd.read_sql(
            "SELECT * FROM production ORDER BY id DESC",
            conn
        )

        st.dataframe(
            df,
            use_container_width=True
        )

        excel_file = create_excel(
            df,
            "Production Report"
        )

        st.download_button(
            "⬇ Download Excel Report",
            excel_file,
            "production_report.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # =================================================
    # KARIGAR LEDGER
    # =================================================

    elif report_type == "Karigar Ledger":

        df = pd.read_sql(
            """
            SELECT
                karigar_name,
                SUM(amount) as total_amount
            FROM production
            GROUP BY karigar_name
            """,
            conn
        )

        st.dataframe(
            df,
            use_container_width=True
        )

        excel_file = create_excel(
            df,
            "Karigar Ledger"
        )

        st.download_button(
            "⬇ Download Excel Report",
            excel_file,
            "karigar_ledger.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # =================================================
    # VENDOR BILLS
    # =================================================

    elif report_type == "Vendor Bills":

        df = pd.read_sql(
            "SELECT * FROM vendor_bills ORDER BY id DESC",
            conn
        )

        st.dataframe(
            df,
            use_container_width=True
        )

        excel_file = create_excel(
            df,
            "Vendor Bills"
        )

        st.download_button(
            "⬇ Download Excel Report",
            excel_file,
            "vendor_bills.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # =================================================
    # VENDOR PAYMENTS
    # =================================================

    elif report_type == "Vendor Payments":

        df = pd.read_sql(
            "SELECT * FROM vendor_payments ORDER BY id DESC",
            conn
        )

        st.dataframe(
            df,
            use_container_width=True
        )

        excel_file = create_excel(
            df,
            "Vendor Payments"
        )

        st.download_button(
            "⬇ Download Excel Report",
            excel_file,
            "vendor_payments.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # =================================================
    # STAFF SALARY
    # =================================================

    elif report_type == "Staff Salary":

        df = pd.read_sql(
            "SELECT * FROM staff_salary ORDER BY id DESC",
            conn
        )

        st.dataframe(
            df,
            use_container_width=True
        )

        excel_file = create_excel(
            df,
            "Staff Salary"
        )

        st.download_button(
            "⬇ Download Excel Report",
            excel_file,
            "staff_salary.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # =================================================
    # STAFF ADVANCE
    # =================================================

    elif report_type == "Staff Advance":

        df = pd.read_sql(
            "SELECT * FROM staff_advances ORDER BY id DESC",
            conn
        )

        st.dataframe(
            df,
            use_container_width=True
        )

        excel_file = create_excel(
            df,
            "Staff Advance"
        )

        st.download_button(
            "⬇ Download Excel Report",
            excel_file,
            "staff_advance.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # ==========================================
# USER MANAGEMENT
# ==========================================

if menu == "User Management":

    st.header("👥 User Management")

    st.subheader("➕ Create New User")

    new_username = st.text_input(
        "Username",
        key="new_user_name"
    )

    new_password = st.text_input(
        "Password",
        type="password",
        key="new_user_password"
    )

    new_role = st.selectbox(
        "Select Role",
        [
            "Admin",
            "Accountant",
            "Cutting",
            "Packing"
        ],
        key="new_user_role"
    )

    if st.button(
        "Create User",
        key="create_user_btn"
    ):

        try:

            cursor.execute(
                """
                INSERT INTO users
                (
                    username,
                    password,
                    role
                )
                VALUES(?,?,?)
                """,
                (
                    new_username,
                    new_password,
                    new_role
                )
            )

            conn.commit()

            st.success(
                "User Created Successfully"
            )

        except:

            st.error(
                "Username Already Exists"
            )

    st.divider()

    # ======================================
    # USER LIST
    # ======================================

    st.subheader("📋 Existing Users")

    user_data = pd.read_sql(
        """
        SELECT

            id as 'ID',

            username as 'Username',

            role as 'Role'

        FROM users

        ORDER BY id DESC
        """,
        conn
    )

    st.dataframe(
        user_data,
        use_container_width=True
    )

    st.divider()

    # ======================================
    # DELETE USER
    # ======================================

    st.subheader("🗑️ Delete User")

    delete_user_id = st.number_input(
        "Enter User ID",
        min_value=1,
        key="delete_user_id"
    )

    if st.button(
        "Delete User",
        key="delete_user_btn"
    ):

        cursor.execute(
            """
            DELETE FROM users
            WHERE id=?
            """,
            (delete_user_id,)
        )

        conn.commit()

        st.success(
            "User Deleted Successfully"
        )
       # =========================================================
# STAFF SALARY MODULE
# =========================================================

elif menu == "Staff Salary":

    st.header("💰 Staff Salary")

    # DATE
    entry_date = st.date_input("Date")

    # STAFF DROPDOWN
    staff_df = pd.read_sql(
        "SELECT staff_name FROM staff_master ORDER BY staff_name",
        conn
    )

    staff_name = st.selectbox(
        "Staff Name",
        staff_df["staff_name"]
    )

    # MONTHLY SALARY
    monthly_salary = st.number_input(
        "Monthly Salary",
        min_value=0.0
    )

    # TOTAL ADVANCE
    advance_data = pd.read_sql(
        f"""
        SELECT SUM(amount) as total_advance
        FROM staff_advances
        WHERE staff_name='{staff_name}'
        """,
        conn
    )

    total_advance = advance_data.iloc[0]["total_advance"]

    if total_advance is None:
        total_advance = 0

    # FINAL SALARY
    final_salary = monthly_salary - total_advance

    st.info(f"Advance = ₹ {total_advance}")

    st.success(f"Final Salary = ₹ {final_salary}")

    # SAVE SALARY
    if st.button("Save Salary"):

        cursor.execute(
            """
            INSERT INTO staff_salary (
                entry_date,
                staff_name,
                monthly_salary,
                advance,
                final_salary,
                payment_status
            )

            VALUES (?, ?, ?, ?, ?, ?)
            """,

            (
                str(entry_date),
                staff_name,
                monthly_salary,
                total_advance,
                final_salary,
                "Pending"
            )
        )

        conn.commit()

        st.success("Salary Saved Successfully")

        st.rerun()


    # =====================================================
    # SALARY HISTORY
    # =====================================================

    st.divider()

    st.subheader("📜 Salary History")

    salary_history = pd.read_sql(
        """
        SELECT *
        FROM staff_salary
        ORDER BY id DESC
        """,
        conn
    )

    if len(salary_history) > 0:

        for index, row in salary_history.iterrows():

            with st.expander(
                f"{row['staff_name']} | ₹ {row['final_salary']} | {row['payment_status']}"
            ):

                st.write(f"Date: {row['entry_date']}")
                st.write(f"Salary: ₹ {row['monthly_salary']}")
                st.write(f"Advance: ₹ {row['advance']}")
                st.write(f"Final Salary: ₹ {row['final_salary']}")

                col1, col2, col3 = st.columns(3)

                # PAID BUTTON
                with col1:

                    if row["payment_status"] == "Pending":

                        if st.button(
                            "Mark Paid",
                            key=f"paid_{row['id']}"
                        ):

                            cursor.execute(
                                """
                                UPDATE staff_salary
                                SET payment_status='Paid'
                                WHERE id=?
                                """,
                                (row["id"],)
                            )

                            conn.commit()

                            st.rerun()

                # DELETE BUTTON
                with col2:

                    if st.button(
                        "Delete",
                        key=f"delete_salary_{row['id']}"
                    ):

                        cursor.execute(
                            "DELETE FROM staff_salary WHERE id=?",
                            (row["id"],)
                        )

                        conn.commit()

                        st.success("Deleted Successfully")

                        st.rerun()
   # =========================================================
# STAFF ADVANCE MODULE
# =========================================================

elif menu == "Staff Advance":

    st.header("💵 Staff Advance")

    # DATE
    entry_date = st.date_input("Date")

    # STAFF DROPDOWN
    staff_df = pd.read_sql(
        "SELECT staff_name FROM staff_master ORDER BY staff_name",
        conn
    )

    staff_name = st.selectbox(
        "Staff Name",
        staff_df["staff_name"],
        key="advance_staff"
    )

    # ADVANCE AMOUNT
    amount = st.number_input(
        "Advance Amount",
        min_value=0.0
    )

    # REMARKS
    remarks = st.text_input("Remarks")

    # SAVE ADVANCE
    if st.button("Save Advance"):

        cursor.execute(
            """
            INSERT INTO staff_advances (
                entry_date,
                staff_name,
                amount,
                remarks
            )

            VALUES (?, ?, ?, ?)
            """,

            (
                str(entry_date),
                staff_name,
                amount,
                remarks
            )
        )

        conn.commit()

        st.success("Advance Saved Successfully")

        st.rerun()


    # =====================================================
    # ADVANCE HISTORY
    # =====================================================

    st.divider()

    st.subheader("📜 Advance History")

    advance_history = pd.read_sql(
        """
        SELECT *
        FROM staff_advances
        ORDER BY id DESC
        """,
        conn
    )

    if len(advance_history) > 0:

        for index, row in advance_history.iterrows():

            with st.expander(
                f"{row['staff_name']} | ₹ {row['amount']}"
            ):

                st.write(f"Date: {row['entry_date']}")
                st.write(f"Amount: ₹ {row['amount']}")
                st.write(f"Remarks: {row['remarks']}")

                # EDIT SECTION
                new_amount = st.number_input(
                    "Edit Amount",
                    value=float(row["amount"]),
                    key=f"amount_{row['id']}"
                )

                new_remarks = st.text_input(
                    "Edit Remarks",
                    value=str(row["remarks"]),
                    key=f"remarks_{row['id']}"
                )

                col1, col2 = st.columns(2)

                # UPDATE BUTTON
                with col1:

                    if st.button(
                        "Update",
                        key=f"update_{row['id']}"
                    ):

                        cursor.execute(
                            """
                            UPDATE staff_advances
                            SET amount=?,
                                remarks=?
                            WHERE id=?
                            """,

                            (
                                new_amount,
                                new_remarks,
                                row["id"]
                            )
                        )

                        conn.commit()

                        st.success("Updated Successfully")

                        st.rerun()

                # DELETE BUTTON
                with col2:

                    if st.button(
                        "Delete",
                        key=f"delete_advance_{row['id']}"
                    ):

                        cursor.execute(
                            "DELETE FROM staff_advances WHERE id=?",
                            (row["id"],)
                        )

                        conn.commit()

                        st.success("Deleted Successfully")

                        st.rerun()
        # ==========================================
# STAFF MASTER
# ==========================================

elif menu == "Staff Master":

    st.header("👨 Staff Master")

    # ADD STAFF
    st.subheader("➕ Add Staff")

    new_staff = st.text_input("Staff Name")

    if st.button("Add Staff"):

        if new_staff != "":

            try:
                cursor.execute(
                    "INSERT INTO staff_master (staff_name) VALUES (?)",
                    (new_staff,)
                )

                conn.commit()

                st.success("Staff Added Successfully")

            except:
                st.error("Staff Already Exists")


    st.divider()

    # STAFF LIST
    st.subheader("📋 Staff List")

    staff_df = pd.read_sql(
        "SELECT * FROM staff_master ORDER BY staff_name",
        conn
    )

    if len(staff_df) > 0:

        for index, row in staff_df.iterrows():

            col1, col2 = st.columns([5,1])

            with col1:
                st.write(row["staff_name"])

            with col2:

                if st.button(
                    "Delete",
                    key=f"delete_staff_{row['id']}"
                ):

                    cursor.execute(
                        "DELETE FROM staff_master WHERE id=?",
                        (row["id"],)
                    )

                    conn.commit()

                    st.success("Staff Deleted")

                    st.rerun()
                    # =========================================================
# VENDORS MODULE
# =========================================================

elif menu == "Vendors":

    st.header("🏢 Vendors Management")

    vendor_menu = st.radio(
        "Select Option",
        [
            "Vendor Master",
            "Vendor Bills",
            "Vendor Payments",
            "Vendor Ledger"
        ],
        horizontal=True
    )

    # =====================================================
    # VENDOR MASTER
    # =====================================================

    if vendor_menu == "Vendor Master":

        st.subheader("Vendor Master")

        vendor_name = st.text_input("Vendor Name")

        vendor_type = st.selectbox(
            "Vendor Type",
            [
                "Fabric",
                "Accessories",
                "Packaging",
                "Transport",
                "Other"
            ]
        )

        if st.button("Save Vendor"):

            if vendor_name != "":

                cursor.execute("""
                    INSERT INTO vendors
                    (
                        vendor_name,
                        vendor_type
                    )
                    VALUES (?, ?)
                """, (
                    vendor_name,
                    vendor_type
                ))

                conn.commit()

                st.success("Vendor Saved Successfully")

                st.rerun()

        st.markdown("---")

        st.subheader("📋 Vendor History")

        vendor_df = pd.read_sql(
            """
            SELECT *
            FROM vendors
            ORDER BY id DESC
            """,
            conn
        )

        if not vendor_df.empty:

            vendor_df.insert(0, "Select", False)

            edited_vendor_df = st.data_editor(
                vendor_df,
                use_container_width=True,
                hide_index=True
            )

            selected_rows = edited_vendor_df[
                edited_vendor_df["Select"] == True
            ]

            if not selected_rows.empty:

                selected_vendor_id = int(
                    selected_rows.iloc[0]["id"]
                )

                if st.button("🗑 Delete Selected Vendor"):

                    cursor.execute(
                        """
                        DELETE FROM vendors
                        WHERE id=?
                        """,
                        (selected_vendor_id,)
                    )

                    conn.commit()

                    st.success("Vendor Deleted Successfully")

                    st.rerun()

    # =====================================================
    # VENDOR BILLS
    # =====================================================

    elif vendor_menu == "Vendor Bills":

        st.subheader("Vendor Bills")

        vendor_df = pd.read_sql(
            """
            SELECT vendor_name
            FROM vendors
            ORDER BY vendor_name
            """,
            conn
        )

        vendor_list = vendor_df["vendor_name"].tolist()

        selected_vendor = st.selectbox(
            "Select Vendor",
            vendor_list
        )

        entry_date = st.date_input("Bill Date")

        bill_no = st.text_input("Bill Number")

        bill_amount = st.number_input(
            "Bill Amount",
            min_value=0.0
        )

        if st.button("Save Bill"):

            cursor.execute("""
                INSERT INTO vendor_bills
                (
                    entry_date,
                    vendor_name,
                    bill_no,
                    bill_amount
                )
                VALUES (?, ?, ?, ?)
            """, (
                str(entry_date),
                selected_vendor,
                bill_no,
                bill_amount
            ))

            conn.commit()

            st.success("Bill Saved Successfully")

            st.rerun()

        st.markdown("---")

        st.subheader("📄 Bill History")

        bill_df = pd.read_sql(
            """
            SELECT
                id,
                entry_date,
                vendor_name,
                bill_no,
                bill_amount
            FROM vendor_bills
            WHERE vendor_name=?
            ORDER BY id DESC
            """,
            conn,
            params=(selected_vendor,)
        )

        if not bill_df.empty:

            bill_df.insert(0, "Select", False)

            edited_bill_df = st.data_editor(
                bill_df,
                use_container_width=True,
                hide_index=True
            )

            selected_rows = edited_bill_df[
                edited_bill_df["Select"] == True
            ]

            if not selected_rows.empty:

                selected_bill_id = int(
                    selected_rows.iloc[0]["id"]
                )

                if st.button("🗑 Delete Selected Bill"):

                    cursor.execute(
                        """
                        DELETE FROM vendor_bills
                        WHERE id=?
                        """,
                        (selected_bill_id,)
                    )

                    conn.commit()

                    st.success("Bill Deleted Successfully")

                    st.rerun()

    # =====================================================
    # VENDOR PAYMENTS
    # =====================================================

    elif vendor_menu == "Vendor Payments":

        st.subheader("Vendor Payments")

        vendor_df = pd.read_sql(
            """
            SELECT vendor_name
            FROM vendors
            ORDER BY vendor_name
            """,
            conn
        )

        vendor_list = vendor_df["vendor_name"].tolist()

        selected_vendor = st.selectbox(
            "Select Vendor",
            vendor_list,
            key="payment_vendor"
        )

        entry_date = st.date_input("Payment Date")

        payment_amount = st.number_input(
            "Payment Amount",
            min_value=0.0
        )

        payment_mode = st.selectbox(
            "Payment Mode",
            [
                "Cash",
                "Bank",
                "UPI"
            ]
        )

        remarks = st.text_input("Remarks")

        if st.button("Save Payment"):

            cursor.execute("""
                INSERT INTO vendor_payments
                (
                    entry_date,
                    vendor_name,
                    payment_amount,
                    payment_mode,
                    remarks
                )
                VALUES (?, ?, ?, ?, ?)
            """, (
                str(entry_date),
                selected_vendor,
                payment_amount,
                payment_mode,
                remarks
            ))

            conn.commit()

            st.success("Payment Saved Successfully")

            st.rerun()

        st.markdown("---")

        st.subheader("💰 Payment History")

        payment_df = pd.read_sql(
            """
            SELECT
                id,
                entry_date,
                vendor_name,
                payment_amount,
                payment_mode,
                remarks
            FROM vendor_payments
            WHERE vendor_name=?
            ORDER BY id DESC
            """,
            conn,
            params=(selected_vendor,)
        )

        if not payment_df.empty:

            payment_df.insert(0, "Select", False)

            edited_payment_df = st.data_editor(
                payment_df,
                use_container_width=True,
                hide_index=True
            )

            selected_rows = edited_payment_df[
                edited_payment_df["Select"] == True
            ]

            if not selected_rows.empty:

                selected_payment_id = int(
                    selected_rows.iloc[0]["id"]
                )

                if st.button("🗑 Delete Selected Payment"):

                    cursor.execute(
                        """
                        DELETE FROM vendor_payments
                        WHERE id=?
                        """,
                        (selected_payment_id,)
                    )

                    conn.commit()

                    st.success("Payment Deleted Successfully")

                    st.rerun()

    # =====================================================
    # VENDOR LEDGER
    # =====================================================

    elif vendor_menu == "Vendor Ledger":

        st.subheader("📒 Vendor Ledger")

        ledger_df = pd.read_sql(
            """
            SELECT
                v.vendor_name,

                IFNULL(b.total_bill, 0) as total_bill,

                IFNULL(p.total_paid, 0) as total_paid,

                (
                    IFNULL(b.total_bill, 0)
                    -
                    IFNULL(p.total_paid, 0)
                ) as pending

            FROM vendors v

            LEFT JOIN
            (
                SELECT
                    vendor_name,
                    SUM(bill_amount) as total_bill
                FROM vendor_bills
                GROUP BY vendor_name
            ) b

            ON v.vendor_name = b.vendor_name

            LEFT JOIN
            (
                SELECT
                    vendor_name,
                    SUM(payment_amount) as total_paid
                FROM vendor_payments
                GROUP BY vendor_name
            ) p

            ON v.vendor_name = p.vendor_name

            ORDER BY v.vendor_name
            """,
            conn
        )

        st.dataframe(
            ledger_df,
            use_container_width=True
        )